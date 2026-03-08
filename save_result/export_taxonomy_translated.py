#!/usr/bin/env python
"""
Export integrated taxonomy structure from all dimensions to Excel (Korean Version with Translation)
"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from openai import OpenAI
from tqdm import tqdm
import sys
import argparse
from dotenv import load_dotenv
sys.path.insert(0, str(Path(__file__).parent))
from config_utils import get_dimensions, get_dimension_names_korean


def translate_to_korean(text, client, description=False):
    """Translate English text to Korean using GPT API"""
    if not text or text == 'None':
        return ''
    
    try:
        if description:
            system_prompt = "You are a professional translator specializing in steel manufacturing and metallurgy. Translate the following technical description to natural Korean. Keep technical terms accurate."
        else:
            system_prompt = "You are a professional translator specializing in steel manufacturing and metallurgy. Translate the following technical term to natural Korean. Keep it concise."
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": text}
            ],
            temperature=0.3,
            max_tokens=500 if description else 100
        )
        
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Translation error: {e}")
        return text  # Return original if translation fails

def parse_taxonomy_tree(node, dimension, client, parent_path="", level=0, rows=None, cache=None, korean_names=None):
    """Recursively parse taxonomy tree and create table rows with translation"""
    if rows is None:
        rows = []
    if cache is None:
        cache = {}
    if korean_names is None:
        korean_names = get_dimension_names_korean()
    
    label = node.get('label', '')
    description = node.get('description', '')
    node_level = node.get('level', level)
    source = node.get('source', 'Initial')
    paper_count = len(node.get('paper_ids', []))
    
    # For Level 0 nodes, use dimension name as label and set special path format
    if node_level == 0:
        display_label = dimension  # Use dimension name for Level 0
        korean_dim = korean_names.get(dimension, dimension)
        current_path = f"{dimension}/{korean_dim}"  # "dimension/한글명" format
        label_kr = korean_dim  # Use korean dimension name for translation
    else:
        display_label = label
        korean_dim = korean_names.get(dimension, dimension)
        current_path = f"{parent_path}/{label}" if parent_path else label
        # Translate label
        if label in cache:
            label_kr = cache[label]
        else:
            label_kr = translate_to_korean(label, client, description=False)
            cache[label] = label_kr
    
    # Translate description
    if description:
        desc_key = f"desc_{description[:50]}"  # Cache key
        if desc_key in cache:
            description_kr = cache[desc_key]
        else:
            description_kr = translate_to_korean(description, client, description=True)
            cache[desc_key] = description_kr
    else:
        description_kr = ''
    
    # Get dimension korean name
    dimension_names_korean = korean_names
    
    # Add current node
    rows.append({
        '차원': dimension_names_korean.get(dimension, dimension),
        '차원_영문': dimension,
        '레벨': node_level,
        '기술명': display_label,
        '기술명_한글': label_kr,
        '전체_경로': current_path,
        '설명': description if description else '',
        '설명_한글': description_kr,
        '출처': source,
        '논문_수': paper_count,
        '계층_구조': ('  ' * node_level + display_label) if node_level > 0 else dimension
    })
    
    # Recursively process children
    children = node.get('children', [])
    if isinstance(children, dict):
        for child_label, child_node in children.items():
            parse_taxonomy_tree(child_node, dimension, client, current_path, node_level + 1, rows, cache, korean_names)
    elif isinstance(children, list):
        for child_node in children:
            parse_taxonomy_tree(child_node, dimension, client, current_path, node_level + 1, rows, cache, korean_names)
    
    return rows

def load_all_taxonomies(data_dir, client):
    """Load all taxonomy JSON files and translate"""
    data_dir = Path(data_dir)
    
    # Use get_dimension_names_korean() directly to get web interface korean names
    korean_names = get_dimension_names_korean()
    
    all_rows = []
    dim_summaries = []
    cache = {}  # Translation cache

    # Scan for all final_taxo_*.json files directly
    json_files = sorted(data_dir.glob("final_taxo_*.json"))
    if not json_files:
        print(f"Warning: No final_taxo_*.json files found in {data_dir}")
        return all_rows, dim_summaries

    for json_file in json_files:
        dim = json_file.stem[len("final_taxo_"):]
        print(f"\nLoading and translating {json_file.name}  (dimension: {dim})...")
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                taxonomy = json.load(f)
        except Exception as e:
            print(f"  - Error reading {json_file.name}: {e}")
            continue

        # Parse tree with translation and korean names
        rows = parse_taxonomy_tree(taxonomy, dim, client, cache=cache, korean_names=korean_names)
        all_rows.extend(rows)

        # Summary statistics
        total_nodes = len(rows)
        total_papers = sum(r['논문_수'] for r in rows)
        max_level = max(r['레벨'] for r in rows) if rows else 0

        dim_summaries.append({
            '차원': korean_names.get(dim, dim),
            '총 노드 수': total_nodes,
            '총 논문 수': total_papers,
            '최대 깊이': max_level
        })

        print(f"  ✓ {total_nodes}개 노드, {total_papers}개 논문, 최대 깊이: {max_level}")

    return all_rows, dim_summaries

def create_excel_report(output_file, all_rows, dim_summaries):
    """Create Excel file with multiple sheets"""
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "요약"
    
    summary_df = pd.DataFrame(dim_summaries)
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True, name='맑은 고딕')
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            else:
                cell.font = Font(name='맑은 고딕')
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = max(len(str(cell.value or "")) * 1.5 for cell in column)
        ws_summary.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 2: All Dimensions (Integrated)
    ws_all = wb.create_sheet("전체 차원")
    
    df_all = pd.DataFrame(all_rows)
    # Reorder columns - show Korean first
    df_all = df_all[['차원', '차원_영문', '레벨', '계층_구조', '기술명', '기술명_한글', '설명', '설명_한글', '출처', '논문_수', '전체_경로']]
    
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True, name='맑은 고딕')
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            else:
                cell.font = Font(name='맑은 고딕')
                # Color code by level
                level = df_all.iloc[r_idx - 2]['레벨']
                if level == 0:
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    cell.font = Font(bold=True, name='맑은 고딕')
                elif level == 1:
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_all.columns:
        max_length = max(len(str(cell.value or "")) * 1.3 for cell in column)
        ws_all.column_dimensions[column[0].column_letter].width = min(max_length + 2, 100)
    
    # Sheets 3-N: Individual Dimensions (derived from actual data)
    df_all_temp = pd.DataFrame(all_rows)
    actual_dims = sorted(df_all_temp['차원'].unique()) if len(all_rows) > 0 else []
    dimensions_list = [(dim, dim) for dim in actual_dims]
    
    for dim_korean, dim_english in dimensions_list:
        ws_dim = wb.create_sheet(dim_korean[:31])  # Excel sheet name limit
        
        df_dim = df_all[df_all['차원'] == dim_korean].copy()
        df_dim = df_dim[['레벨', '계층_구조', '기술명', '기술명_한글', '설명', '설명_한글', '출처', '논문_수', '전체_경로']]
        
        for r_idx, row in enumerate(dataframe_to_rows(df_dim, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_dim.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header
                    cell.font = Font(bold=True, name='맑은 고딕')
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                else:
                    cell.font = Font(name='맑은 고딕')
                    # Color code by level
                    if len(df_dim) > 0:
                        level = df_dim.iloc[r_idx - 2]['레벨']
                        if level == 0:
                            cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                            cell.font = Font(bold=True, name='맑은 고딕')
                        elif level == 1:
                            cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws_dim.columns:
            max_length = max(len(str(cell.value or "")) * 1.3 for cell in column)
            ws_dim.column_dimensions[column[0].column_letter].width = min(max_length + 2, 100)
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ 한글 번역 엑셀 파일 저장 완료: {output_file}")
    
    return output_file

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Export integrated taxonomy structure from all dimensions to Excel (Korean Version with Translation)')
    parser.add_argument('--data-dir', default='/root/go_taxoadapt/datasets/posco', 
                       help='Directory containing taxonomy JSON files (default: /root/go_taxoadapt/datasets/posco)')
    parser.add_argument('--output', help='Output Excel file path (default: saves to data directory as integrated_taxonomy_structure_korean_translated.xlsx)')
    
    args = parser.parse_args()
    
    # Setup paths
    data_dir = Path(args.data_dir)
    
    if not data_dir.exists():
        print(f"Error: Data directory not found: {data_dir}")
        return
    
    # Generate output path if not specified
    if args.output:
        output_file = Path(args.output)
    else:
        output_file = data_dir / 'integrated_taxonomy_structure_korean_translated.xlsx'
    
    # Load environment variables from .env file
    load_dotenv()
    
    # Initialize OpenAI client
    api_key = os.getenv("OPENAI_API_KEY")

    if not api_key:
        print("Error: OPENAI_API_KEY not found in environment variables.")
        print("Please make sure your .env file contains OPENAI_API_KEY=your_api_key")
        return
    
    client = OpenAI(api_key=api_key)
    
    print("=" * 70)
    print("통합 기술 분류 체계 엑셀 생성 (한글 번역 버전)")
    print("=" * 70)
    print(f"데이터 디렉토리: {data_dir}")
    print(f"출력 파일: {output_file}")
    
    # Load all taxonomies
    print("\n[1/2] 기술 분류 체계 로딩 및 번역 중...")
    print("(GPT API를 사용하여 기술명과 설명을 한글로 번역합니다)")
    all_rows, dim_summaries = load_all_taxonomies(data_dir, client)
    
    # Create Excel report
    print("\n[2/2] 엑셀 리포트 생성 중...")
    output_file = create_excel_report(output_file, all_rows, dim_summaries)
    
    # Print final statistics
    print("\n" + "=" * 70)
    print("생성 완료!")
    print("=" * 70)
    print(f"전체 차원의 총 노드 수: {len(all_rows)}")
    print(f"출력 파일: {output_file}")
    print("\n생성된 시트:")
    print("  1. 요약 - 전체 차원 개요")
    print("  2. 전체 차원 - 통합 뷰 (한글/영문 병기)")
    print("  3-6. 개별 차원별 시트 (한글/영문 병기)")
    print("\n※ 기술명과 설명이 GPT를 통해 한글로 번역되었습니다.")

if __name__ == "__main__":
    main()
