#!/usr/bin/env python
"""
Export integrated taxonomy structure from all dimensions to Excel
"""
import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import argparse
import os
sys.path.insert(0, str(Path(__file__).parent))
from config_utils import get_dimensions, get_dimension_names_korean

def parse_taxonomy_tree(node, dimension, parent_path="", level=0, rows=None, korean_names=None):
    """Recursively parse taxonomy tree and create table rows"""
    if rows is None:
        rows = []
    if korean_names is None:
        korean_names = get_dimension_names_korean()
    
    label = node.get('label', '')
    description = node.get('description', '')
    node_level = node.get('level', level)
    source = node.get('source', 'Initial')
    paper_count = len(node.get('paper_ids', []))
    
    # For Level 0 nodes, use dimension name as label and set special path format
    if node_level == 0:
        display_label = dimension  # Use dimension name for Level 0
        korean_dim = korean_names.get(dimension, dimension)
        current_path = f"{dimension}/{korean_dim}"  # "dimension/한글명" format
        indent_text = dimension  # Use dimension name for indent
    else:
        display_label = label
        korean_dim = korean_names.get(dimension, dimension)
        current_path = f"{parent_path}/{label}" if parent_path else label
        indent_text = '  ' * node_level + label
    
    # Add current node
    rows.append({
        'Dimension': dimension,
        'Dimension_Korean': korean_dim,
        'Level': node_level,
        'Label': display_label,
        'Full_Path': current_path,
        'Description': description,
        'Source': source,
        'Paper_Count': paper_count,
        'Indent': indent_text  # For visual hierarchy
    })
    
    # Recursively process children
    children = node.get('children', [])
    if isinstance(children, dict):
        for child_label, child_node in children.items():
            parse_taxonomy_tree(child_node, dimension, current_path, node_level + 1, rows, korean_names)
    elif isinstance(children, list):
        for child_node in children:
            parse_taxonomy_tree(child_node, dimension, current_path, node_level + 1, rows, korean_names)
    
    return rows

def load_all_taxonomies(data_dir):
    """Load all taxonomy JSON files by scanning directory directly"""
    data_dir = Path(data_dir)
    all_rows = []
    dim_summaries = []
    
    # Use get_dimension_names_korean() directly to get web interface korean names
    korean_names = get_dimension_names_korean()

    # Scan for all final_taxo_*.json files directly
    json_files = sorted(data_dir.glob("final_taxo_*.json"))
    if not json_files:
        print(f"Warning: No final_taxo_*.json files found in {data_dir}")
        return all_rows, dim_summaries

    for json_file in json_files:
        dim = json_file.stem[len("final_taxo_"):]
        print(f"Loading {json_file.name}  (dimension: {dim})...")
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                taxonomy = json.load(f)
        except Exception as e:
            print(f"  - Error reading {json_file.name}: {e}")
            continue

        # Parse tree with korean names
        rows = parse_taxonomy_tree(taxonomy, dim, korean_names=korean_names)
        all_rows.extend(rows)

        # Summary statistics
        total_nodes = len(rows)
        total_papers = sum(r['Paper_Count'] for r in rows)
        max_level = max(r['Level'] for r in rows) if rows else 0

        dim_summaries.append({
            'Dimension': dim,
            'Dimension_Korean': korean_names.get(dim, dim),
            'Total_Nodes': total_nodes,
            'Total_Papers': total_papers,
            'Max_Depth': max_level
        })

        print(f"  - {total_nodes} nodes, {total_papers} papers, max depth: {max_level}")

    return all_rows, dim_summaries

def create_excel_report(output_file, all_rows, dim_summaries):
    """Create Excel file with multiple sheets"""
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_df = pd.DataFrame(dim_summaries)
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_summary.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 2: All Dimensions (Integrated)
    ws_all = wb.create_sheet("All_Dimensions")
    
    df_all = pd.DataFrame(all_rows)
    # Reorder columns
    df_all = df_all[['Dimension', 'Dimension_Korean', 'Level', 'Indent', 'Label', 'Description', 'Source', 'Paper_Count', 'Full_Path']]
    
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            else:
                # Color code by level
                level = df_all.iloc[r_idx - 2]['Level']
                if level == 0:
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    cell.font = Font(bold=True)
                elif level == 1:
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_all.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws_all.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    
    # Sheets 3-N: Individual Dimensions (derived from actual data)
    actual_dims = sorted(df_all['Dimension'].unique()) if len(df_all) > 0 else []
    korean_names = get_dimension_names_korean()
    dimensions = [(dim, korean_names.get(dim, dim)[:30]) for dim in actual_dims]
    
    for dim_full, dim_short in dimensions:
        ws_dim = wb.create_sheet(dim_short)
        
        df_dim = df_all[df_all['Dimension'] == dim_full].copy()
        df_dim = df_dim[['Level', 'Indent', 'Label', 'Description', 'Source', 'Paper_Count', 'Full_Path']]
        
        for r_idx, row in enumerate(dataframe_to_rows(df_dim, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_dim.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                else:
                    # Color code by level
                    level = df_dim.iloc[r_idx - 2]['Level']
                    if level == 0:
                        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                        cell.font = Font(bold=True)
                    elif level == 1:
                        cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws_dim.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws_dim.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel file saved: {output_file}")
    
    return output_file

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Export integrated taxonomy structure from all dimensions to Excel')
    parser.add_argument('--data-dir', default='./datasets/posco', 
                       help='Directory containing taxonomy JSON files (default: ./datasets/posco)')
    parser.add_argument('--output', help='Output Excel file path (default: saves to data directory as integrated_taxonomy_structure.xlsx)')
    
    args = parser.parse_args()
    
    # Setup paths
    data_dir = Path(args.data_dir)
    
    if not data_dir.exists():
        print(f"Error: Data directory not found: {data_dir}")
        return
    
    # Generate output path if not specified (save in main directory)
    if args.output:
        output_file = Path(args.output)
    else:
        main_dir = Path(__file__).parent.parent  # go_taxoadapt main directory
        output_file = main_dir / 'electrical_steel_taxonomy_structure.xlsx'
    
    print("=" * 70)
    print("Electrical Steel Taxonomy Structure Export")
    print("=" * 70)
    print(f"Data directory: {data_dir}")
    print(f"Output file: {output_file}")
    
    # Load all taxonomies
    print("\n[1/2] Loading taxonomy structures...")
    all_rows, dim_summaries = load_all_taxonomies(data_dir)
    
    # Create Excel report
    print("\n[2/2] Creating Excel report...")
    output_file = create_excel_report(output_file, all_rows, dim_summaries)
    
    # Print final statistics
    print("\n" + "=" * 70)
    print("Export Complete!")
    print("=" * 70)
    print(f"Total nodes across all dimensions: {len(all_rows)}")
    print(f"Output file: {output_file}")
    print("\nSheets created:")
    print("  1. Summary - Overview of all electrical steel dimensions")
    print("  2. All_Dimensions - Integrated view of all dimensions")
    print("  3-11. Individual dimension sheets (Korean names)")

if __name__ == "__main__":
    main()
